import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl import Workbook
import re

# https://kess.kedi.re.kr/index
# https://www.vworld.kr/dev/v4dv_geocoderguide2_s001.do

# filePath = r"27. 전국대학교 위치 시각화\학교주소좌표.xlsx"
# df_from_excel = pd.read_excel(filePath, engine="openpyxl")

# df_from_excel.columns = df_from_excel.loc[4].tolist()

# df_from_excel = df_from_excel.drop(index=list(range(0, 5)))

# print(df_from_excel.head())
# print(df_from_excel["학교명"].values)
# print(df_from_excel["주소"].values)

apiurl = "https://api.vworld.kr/req/address?"
params = "service=address&request=getcoord&version=2.0&crs=epsg:4326&refine=true&simple=-false&format=json&type="
road_type = "ROAD"
road_type2 = "PARCEL"
address = "&address="
keys = "&key="
primary_key = "EA54B124-084B-3BF4-866B-5B0485B5A2D2"


def request_geo(road):
    page = requests.get(
        apiurl + params + road_type + address + road + keys + primary_key
    )
    json_data = page.json()
    if json_data["response"]["status"] == "OK":
        x = json_data["response"]["result"]["point"]["x"]
        y = json_data["response"]["result"]["point"]["y"]
        return x, y
    else:
        x = 0
        y = 0
        return x, y

x, y = request_geo("경기도 시흥시 산기대학로 237 (정왕동, 한국산업기술대학교)")

print(f'값: {x}')
print(f'값: {y}')

# try:
#     wb = load_workbook(r"27. 전국대학교 위치 시각화\학교주소좌표.xlsx", data_only=True)
#     sheet = wb.active
# except:
#     wb = Workbook()
#     sheet = wb.active

# university_list = df_from_excel["학교명"].to_list()
# address_list = df_from_excel["주소"].to_list()
# print(address_list)
# for num, value in enumerate(address_list):
#     addr = re.sub(r'\([^)]*\)', '', value)
#     print(addr)
#     x, y = request_geo(addr)
#     sheet.append([university_list[num], addr, x, y])

# wb.save(r"27. 전국대학교 위치 시각화\학교주소좌표.xlsx")
